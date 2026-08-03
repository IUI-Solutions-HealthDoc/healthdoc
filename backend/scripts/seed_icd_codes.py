"""Seed/refresh the local icd_codes catalog (run after migration 0007).

Usage:
  python scripts/seed_icd_codes.py icd10 path/to/icd10_codes.csv
      CSV columns: code,title
  python scripts/seed_icd_codes.py icd11 path/to/who_mms_simpletabulation.xlsx
      Official WHO ICD-11 MMS 'SimpleTabulation' release file
      (https://icd.who.int/ -> ICD-11 -> Downloads). Loads code, title, URI,
      and postcoordination flag. Idempotent upsert on (version, code).
"""
import asyncio
import csv
import sys

from sqlalchemy import text

from app.common.db import SessionLocal

UPSERT = text("""
    INSERT INTO icd_codes (version, code, title, icd_uri, is_postcoordinable)
    VALUES (:version, :code, :title, :icd_uri, :pc)
    ON CONFLICT (version, code) DO UPDATE
      SET title = EXCLUDED.title,
          icd_uri = EXCLUDED.icd_uri,
          is_postcoordinable = EXCLUDED.is_postcoordinable,
          updated_at = now()
""")


def rows_icd10(path: str):
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            yield {"version": "icd10", "code": r["code"].strip(),
                   "title": r["title"].strip(), "icd_uri": None, "pc": False}


def rows_icd11(path: str):
    from openpyxl import load_workbook  # pip install openpyxl --break-system-packages
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
    idx = {name: header.index(name) for name in header}
    code_col = idx.get("Code"); title_col = idx.get("Title")
    uri_col = idx.get("Linearization (release) URI", idx.get("Foundation URI"))
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[code_col]
        if not code:            # grouping rows have no code
            continue
        title = str(row[title_col] or "").lstrip("- ").strip()
        uri = str(row[uri_col]) if uri_col is not None and row[uri_col] else None
        yield {"version": "icd11", "code": str(code).strip(),
               "title": title, "icd_uri": uri, "pc": "&" not in str(code)}


async def main(version: str, path: str) -> None:
    rows = list(rows_icd10(path) if version == "icd10" else rows_icd11(path))
    async with SessionLocal() as session:
        for i in range(0, len(rows), 1000):
            for r in rows[i:i + 1000]:
                await session.execute(UPSERT, r)
            await session.commit()
    print(f"Upserted {len(rows)} {version} codes")


if __name__ == "__main__":
    if len(sys.argv) != 3 or sys.argv[1] not in ("icd10", "icd11"):
        print(__doc__); sys.exit(1)
    asyncio.run(main(sys.argv[1], sys.argv[2]))
